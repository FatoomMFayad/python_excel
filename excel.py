import openpyxl as xl

class Country:
    name = ''
    imports = 0
    exports = 0
    capital = ''


countries = []

file_name = 'C:/Users/Fatoom/Documents/countries.xlsx'
try:
    wb = xl.load_workbook(file_name)
    sheet = wb.active

    # print(sheet.cell(row=1, column=1).value)
    # print(sheet['A2'].value)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        c = Country()
        c.name = row[0]
        c.imports = row[1]
        c.exports = row[2]
        c.capital = row[3]
        countries.append(c)

    for i in countries:
        print(i.name, i.imports, i.exports, i.capital)

    total_imports = sum(i.imports for i in countries)
    total_exports = sum(i.exports for i in countries)

    sheet['B5'] = total_imports
    sheet['C5'] = total_exports
    results_sheet = wb.create_sheet('results')
    results_sheet['A1'] = 'total imports'
    results_sheet['A2'] = total_imports
    results_sheet['B1'] = 'total exports'
    results_sheet['B2'] = total_exports
    wb.save(file_name)

except FileNotFoundError as e:
    print(e)